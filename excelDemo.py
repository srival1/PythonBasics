import openpyxl

book = openpyxl.load_workbook("C:\\Users\\srika\\PycharmProjects\\PythonBasics\\signup_info.xlsx")


#sheet = book.active
sheet = book["Sheet1"]
print(sheet)
cell = sheet.cell(row=2, column=2)
print(cell.value)

sheet.cell(row=1, column=1).value = "testcase#"
print(sheet.cell(row=1, column=1).value)

print(sheet.max_row)
print(sheet.max_column)

print(sheet['A1'].value)
print(sheet['F3'].value)
print("*****************************************************")
for rowi in range(1, sheet.max_row+1):
    for colj in range(1, sheet.max_column+1):
        print(sheet.cell(row=rowi, column=colj).value)

print("*****************************************************")
for rowi in range(1, sheet.max_row+1):
    if sheet.cell(row=rowi, column=1).value == "testcase2":
        for colj in range(1, sheet.max_column+1):
            print(sheet.cell(row=rowi, column=colj).value)

test_dict = {}
print("*****************************************************")
Dict = {}
for i in range(1, sheet.max_row + 1):  # to get rows
    if sheet.cell(row=i, column=1).value == "testcase1":
        for j in range(2, sheet.max_column + 1):  # to get columns
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
print(Dict)
print(type(Dict["dob"].date()))
strdate = Dict["dob"].date().strftime("%m/%d/%Y")
print(strdate)
print(type(strdate))
print("*****************************************************")

import pandas as pd

# convert into dataframe
df = pd.read_excel("C:\\Users\\srika\\PycharmProjects\\PythonBasics\\signup_info.xlsx")

# convert into dictionary
dict = df.to_dict()
print(dict)
# print("*****************************************************")
excel = pd.ExcelFile("C:\\Users\\srika\\PycharmProjects\\PythonBasics\\signup_info.xlsx")
data = excel.parse(excel.sheet_names[0])
print(data.to_dict())
"""